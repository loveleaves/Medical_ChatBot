# -*- coding:utf-8 -*-
import os
import pickle
import random
import codecs
import numpy as np
import pandas as pd
import json
from openpyxl import Workbook
from tqdm import tqdm

from entity_normalization.bm25_retrival import BM25Retrieval
from entity_normalization.config import raw_data_params, data_params, esim_params


class SaveFile:
    def save_to_xlsx(self, data: list, file_name: str, **kwargs):
        wb = Workbook()
        ws = wb.active

        if 'head' in kwargs and kwargs['head']:
            # 只写到6个表头
            head_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
            for i, head in enumerate(kwargs['head']):
                ws[head_list[i] + '1'] = head

        else:
            # default head
            ws['A1'] = '原始词'
            ws['B1'] = '标准词'

        for line in data:
            try:
                ws.append(line)
            except Exception as e:
                # 对于有些含有"\x04"等字符的用例直接删去
                # print(e, line)
                pass

        wb.save(file_name)

    def save_to_txt(self, data: list, file_name: str, **kwargs):
        with open(file_name, 'a', encoding='utf-8') as f:
            for line in data:
                if isinstance(line, list):
                    for item in line[:-1]:
                        f.write(str(item) + "\t")
                    f.write(str(line[-1]) + "\n")

                else:
                    f.write(str(line) + "\n")


class ConvertDType:
    """
    function : 把CHIP-CDN数据处理成MakeData数据格式
    """

    def __init__(self):
        self.raw_data_dir = "./CHIP-CDN/"
        self.raw_data_param = {
            'train_file': self.raw_data_dir + "CHIP-CDN_train.json",
            'val_file': self.raw_data_dir + "CHIP-CDN_dev.json",
            # 'answer_file': self.raw_data_dir + "手术标注词2500.tsv", # not used
            'code_file': self.raw_data_dir + "国际疾病分类 ICD-10北京临床版v601.xlsx"
        }
        self.target_data_params = {
            'train_file': self.raw_data_dir + "train.xlsx",
            'val_file': self.raw_data_dir + "val.xlsx",
            'answer_file': self.raw_data_dir + "answer.xlsx",
            'code_file': self.raw_data_dir + "code.txt"
        }

    def data_process(self, data: str, **kwargs) -> str:
        if 'strip_tag' in kwargs and kwargs['strip_tag']:
            stop_tag = [';', ',', '?', ':', '!', '.', "'", '"',
                        '；', '，', '？', '：', '！', '。', '‘', '“']
            for tag in stop_tag:
                data = data.strip(tag)

        return data

    def convert(self):
        """
        function : convert data type
        """
        # convert train_file
        with open(self.raw_data_param['train_file'], 'r', encoding='utf-8') as f:
            raw_data = json.load(f)
        save_helper = SaveFile()

        train_data = []
        for item in raw_data:
            text = self.data_process(item['text'], strip_tag=True)
            normalized_result = self.data_process(item['normalized_result'], strip_tag=True)
            train_data.append([text, normalized_result])
        save_helper.save_to_xlsx(train_data, self.target_data_params['train_file'])

        # convert val_file
        with open(self.raw_data_param['val_file'], 'r', encoding='utf-8') as f:
            raw_data = json.load(f)

        val_data = []
        for item in raw_data:
            text = self.data_process(item['text'], strip_tag=True)
            normalized_result = self.data_process(item['normalized_result'], strip_tag=True)
            val_data.append([text, normalized_result])
        save_helper.save_to_xlsx(val_data, self.target_data_params['val_file'])

        # generate empty answer.xlsx (Because "手术标注词2500.tsv" file is not used)
        answer_data = []
        save_helper.save_to_xlsx(answer_data, self.target_data_params['answer_file'])

        code_data = []
        raw_data = pd.read_excel(self.raw_data_param['code_file'])
        for index in range(len(raw_data)):
            codes = raw_data.iloc[index, 0].split('+')
            name = raw_data.iloc[index, 1]
            for code in codes:
                if code.strip():
                    code_data.append([code, name])
        save_helper.save_to_txt(code_data, self.target_data_params['code_file'])

    def merge_data(self):
        n7k_RAW_DATASET_DIR = "./yidu-n7k/"
        n7k_raw_data_params = {
            'train_file': n7k_RAW_DATASET_DIR + "train.xlsx",
            'val_file': n7k_RAW_DATASET_DIR + "val.xlsx",
            'answer_file': n7k_RAW_DATASET_DIR + "answer.xlsx",
            'code_file': n7k_RAW_DATASET_DIR + "code.txt"
        }
        save_helper = SaveFile()

        train1_df = pd.read_excel(self.target_data_params['train_file'])
        train2_df = pd.read_excel(n7k_raw_data_params['train_file'])
        data = np.concatenate([train1_df.values, train2_df.values], axis=0)
        save_helper.save_to_xlsx(data.tolist(), raw_data_params['train_file'])

        answer1_df = pd.read_excel(self.target_data_params['answer_file'])
        answer2_df = pd.read_excel(n7k_raw_data_params['answer_file'])
        data = np.concatenate([answer1_df.values, answer2_df.values], axis=0)
        save_helper.save_to_xlsx(data.tolist(), raw_data_params['answer_file'])

        val1_df = pd.read_excel(self.target_data_params['val_file'])
        val2_df = pd.read_excel(n7k_raw_data_params['val_file'])
        data = np.concatenate([val1_df.values, val2_df.values], axis=0)
        save_helper.save_to_xlsx(data.tolist(), raw_data_params['val_file'])

        data = []
        with open(self.target_data_params['code_file'], 'r', encoding='utf8') as f:
            for line in f.readlines():
                code, name = line.strip().split('\t')
                data.append([code, name])
        with open(n7k_raw_data_params['code_file'], 'r', encoding='utf8') as f:
            for line in f.readlines():
                code, name = line.strip().split('\t')
                data.append([code, name])
        save_helper.save_to_txt(data, raw_data_params['code_file'])


class MakeData:
    def gen_training_data(self):
        train_df = pd.read_excel(raw_data_params['train_file'])
        answer_df = pd.read_excel(raw_data_params['answer_file'])
        val_df = pd.read_excel(raw_data_params['val_file'])

        self.bm25Model = BM25Retrieval(raw_data_params['code_file'])
        self.total = 0
        self.error = 0

        train = []
        data = np.concatenate([train_df.values, answer_df.values], axis=0)
        for raw_entity, norm_entity in tqdm(data, desc="数据集生成中"):
            # 需要拆分具有多个实体的样本
            if '+' not in raw_entity and "##" not in norm_entity:
                train.append([raw_entity, norm_entity, 1])
                for neg in self.gen_negative_sample(raw_entity, norm_entity):
                    train.append([raw_entity, neg, 0])
            elif '+' not in raw_entity and "##" in norm_entity:
                for ne in norm_entity.split("##"):
                    train.append([raw_entity, ne, 1])
                    for neg in self.gen_negative_sample(raw_entity, ne):
                        train.append([raw_entity, neg, 0])
            elif '+' in raw_entity and "##" in norm_entity:
                ne_list = norm_entity.split("##")
                re_list = raw_entity.split("+")
                for raw_ent, norm_ent in self.match_positive_sample(re_list, ne_list):
                    train.append([raw_ent, norm_ent, 1])
                    for neg in self.gen_negative_sample(raw_ent, norm_ent):
                        train.append([raw_ent, neg, 0])
            else:
                re_list = raw_entity.split("+")
                for re in re_list:
                    train.append([re, norm_entity, 1])
                    for neg in self.gen_negative_sample(re, norm_entity):
                        train.append([re, neg, 0])

        train = pd.DataFrame(train)
        train.columns = ["sentence1", "sentence2", "label"]
        train.drop_duplicates(subset=["sentence1", "sentence2"], keep='first')

        test = []
        for raw_entity, norm_entity in val_df.values:
            # 需要拆分具有多个实体的样本
            if '+' not in raw_entity and "##" not in norm_entity:
                test.append([raw_entity, norm_entity, 1])

            elif '+' not in raw_entity and "##" in norm_entity:
                for ne in norm_entity.split("##"):
                    test.append([raw_entity, ne, 1])

            elif '+' in raw_entity and "##" in norm_entity:
                ne_list = norm_entity.split("##")
                re_list = raw_entity.split("##")
                for raw_ent, norm_ent in self.match_positive_sample(re_list, ne_list):
                    test.append([raw_ent, norm_ent, 1])
            else:
                re_list = raw_entity.split("##")
                for re in re_list:
                    test.append([re, norm_entity, 1])

        test = pd.DataFrame(test)
        test.columns = ["sentence1", "sentence2", "label"]

        train.to_csv(data_params['train_file'], index=False, encoding="utf8")
        test.to_csv(data_params['test_file'], index=False, encoding="utf8")
        print(train.shape)
        print("bm25 未召回正确规范实体的比例：", self.error / self.total)

        return train, test

    def gen_negative_sample(self, raw, norm):
        """
        function : 生成9个最相关的负样本（因为数据集里只有正样本）
        """
        cand_entity = self.bm25Model.retrieval(raw, 100)
        self.total += 1
        try:
            cand_entity.remove(norm)
        except:
            self.error += 1
        return cand_entity

    def match_positive_sample(self, list1, list2):
        ress = []
        for e1 in list1:
            score = 0
            pos_e = ""
            for e2 in list2:
                s = len(set(e1) & set(e2))
                if s > score:
                    score = s
                    pos_e = e2
            if pos_e != "":
                ress.append([e1, pos_e])

        return ress


def pad_sequences(sequences, maxlen=None, dtype='int32', padding='post',
                  truncating='post', value=0.):
    """
    function :
        把序列长度转变为一样长的，如果设置了maxlen则长度统一为maxlen，如果没有设置则默认取
        最大的长度。填充和截取包括两种方法，post与pre，post指从尾部开始处理，pre指从头部
        开始处理，默认都是从尾部开始。

    params :
        sequences: 序列
        maxlen: int 最大长度
        dtype: 转变后的数据类型
        padding: 填充方法'pre' or 'post'
        truncating: 截取方法'pre' or 'post'
        value: float 填充的值

    Returns:
        x: numpy array 填充后的序列维度为 (number_of_sequences, maxlen)

    """
    lengths = [len(s) for s in sequences]

    nb_samples = len(sequences)
    if maxlen is None:
        maxlen = np.max(lengths)

    x = (np.ones((nb_samples, maxlen)) * value).astype(dtype)
    for idx, s in enumerate(sequences):
        if len(s) == 0:
            continue  # empty list was found
        if truncating == 'pre':
            trunc = s[-maxlen:]
        elif truncating == 'post':
            trunc = s[:maxlen]
        else:
            raise ValueError("Truncating type '%s' not understood" % padding)

        if padding == 'post':
            x[idx, :len(trunc)] = trunc
        elif padding == 'pre':
            x[idx, -len(trunc):] = trunc
        else:
            raise ValueError("Padding type '%s' not understood" % padding)
    return x


def shuffle(*args):
    """
    function : Shuffle 数据
    params : *arrs(数组数据)
    """
    arrs = list(args)
    for i, arr in enumerate(arrs):
        assert len(arrs[0]) == len(arrs[i])
        arrs[i] = np.array(arr)
    p = np.random.permutation(len(arrs[0]))
    return tuple(arr[p] for arr in arrs)


def load_char_vocab():
    if os.path.exists(esim_params['word2id']):
        word2idx, idx2word = pickle.load(open(esim_params['word2id'], "rb"))
    else:
        df = pd.read_csv(data_params['train_file'], encoding="utf8")
        vocab = []
        for ent in df["sentence1"].tolist() + df["sentence2"].tolist():
            try:
                vocab.extend(list(ent))
            except:
                # 数据构造时可能：未去除负样本为空的情况，注意查看数据构造环节的问题
                # assert '' == "a", "数据：‘" +str(ent) + "’有问题，类型：" + str(type(ent))
                continue

        with open(os.path.join(data_params['code_file']), encoding='utf8') as f:
            for line in f.readlines():
                code, name = line.strip().split('\t')
                vocab.extend(list(name))

        vocab = sorted(set(vocab))
        print(len(vocab))
        word2idx = {word: index for index, word in enumerate(vocab, start=2)}
        word2idx["UNK"] = 1
        idx2word = {index: word for word, index in word2idx.items()}
        pickle.dump((word2idx, idx2word), open(esim_params['word2id'], "wb"))

    return word2idx, idx2word


def char_index(p_sentences, h_sentences, maxlen=35):
    word2idx, idx2word = load_char_vocab()

    p_list, h_list = [], []
    for p_sentence, h_sentence in zip(p_sentences, h_sentences):
        p = [word2idx[word.lower()] for word in str(p_sentence) if
             len(word.strip()) > 0 and word.lower() in word2idx.keys()]
        h = [word2idx[word.lower()] for word in str(h_sentence) if
             len(word.strip()) > 0 and word.lower() in word2idx.keys()]

        p_list.append(p)
        h_list.append(h)

    p_list = pad_sequences(p_list, maxlen=maxlen)
    h_list = pad_sequences(h_list, maxlen=maxlen)

    return p_list, h_list

def load_char_embed(feature_size, embed_size):
    """
    function: 加载ESIM模型的Embedding层词向量
    TODO
    """
    pass

def load_char_data(path, data_size=None, maxlen=35):
    df = pd.read_csv(path)
    p = df['sentence1'].values[0:data_size]
    h = df['sentence2'].values[0:data_size]
    label = df['label'].values[0:data_size]

    p, h, label = shuffle(p, h, label)

    # [1,2,3,4,5] [4,1,5,2,0]
    p_c_index, h_c_index = char_index(p, h, maxlen=maxlen)

    return p_c_index, h_c_index, label


def make_code():
    file_path = "./total/diseases.json"
    target_path = "./total/code.txt"

    assert os.path.exists(file_path), "文件不存在！"
    with open(file_path, 'r') as f:
        data = json.load(f)
    first_num = 1
    data = set(data)
    # 不需要添加下面的原词数据
    # df = pd.read_excel("./total/alias.xlsx")
    # for i in range(len(df)):
    #     item = str(df.iloc[i][1])
    #     for disease in item.split("##"):
    #         data.add(disease)

    assert not os.path.exists(target_path), "文件已存在！"
    with open(target_path, 'a', encoding='utf-8') as f:
        for item in data:
            f.write(str(first_num / 100) + "\t" + item + "\n")
            first_num += 1


def convert_diy_data():
    file_path = "./total/alias.xlsx"

    saver = SaveFile()
    saver.save_to_xlsx([], "./total/answer.xlsx")

    data = []
    df = pd.read_excel(file_path)
    for i in range(len(df)):
        # correct the order
        original_text = df.iloc[i][1]
        for original_word in original_text.split("##"):
            data.append([original_word, df.iloc[i][0]])
    random.shuffle(data)

    train_pos = int(0.8 * len(df))
    train_data = data[:train_pos]
    saver.save_to_xlsx(train_data, "./total/train.xlsx")
    val_data = data[train_pos:]
    saver.save_to_xlsx(val_data, "./total/val.xlsx")


def save_negative_sample(data: list):
    """
    function: 保存高于阈值的负样本
    """
    train_data = []
    df = pd.read_csv(data_params['train_file'])
    for i in range(len(df)):
        train_data.append([df.iloc[i][0], df.iloc[i][1], df.iloc[i][2]])
    train_data.extend(data)
    train_df = pd.DataFrame(train_data)
    train_df.columns = ["sentence1", "sentence2", "label"]
    train_df.drop_duplicates(subset=["sentence1", "sentence2"], keep='first')

    train_df.to_csv(data_params['train_extend_file'], index=False, encoding="utf8")

def make_dictsearch_data():
    file_path = "./total/alias.xlsx"

    code_data = []
    dict_data = {}
    df = pd.read_excel(file_path)
    for i in range(len(df)):
        normalized_entity = df.iloc[i][0]
        alias_text = df.iloc[i][1].split("##")
        code_data.extend(alias_text)
        for alias in alias_text:
            if alias not in dict_data.keys():
                dict_data[alias] = [normalized_entity]
            else:
                if normalized_entity not in dict_data[alias]:
                    dict_data[alias].append(normalized_entity)

    with open("./checkpoint/alias_code.txt", 'a', encoding='utf-8') as f:
        for line in code_data:
            f.write(line + "\n")

    with codecs.open("./checkpoint/alias_dict.json", 'w', encoding='utf-8') as f:
        json.dump(dict_data, f, indent=4, ensure_ascii=False)

if __name__ == '__main__':
    # convert CHIP-CDN rawdata to data
    # converter = ConvertDType()
    # converter.convert()
    # converter.merge_data()

    # convert yidu-n7k rawdata to data
    MakeData_entry = MakeData()
    MakeData_entry.gen_training_data()
    # make_dictsearch_data()

    # make_code()
    # convert_diy_data()
