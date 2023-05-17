# coding: utf-8
import os
import urllib.request
import urllib.parse
import threading
import json
import re
import queue
from lxml import etree
from openpyxl import Workbook


class myThread(threading.Thread):
    """
    function: 为线程定义一个函数
    """

    def __init__(self, name, q):
        threading.Thread.__init__(self)
        self.name = name
        self.q = q

    def run(self):
        print("Starting " + self.name)
        handler = CrimeSpider()
        while True:
            try:
                name = self.q.get(timeout=2)
                handler.spider_main(name)
            except:
                break
        print("Exiting " + self.name)


class SaveFile:
    def save_to_xlsx(self, data: list, file_name: str, **kwargs):
        assert not os.path.exists(file_name), "文件已存在！"
        wb = Workbook()
        ws = wb.active

        if 'head' in kwargs and kwargs['head']:
            # 只写到6个表头
            head_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
            for i, head in enumerate(kwargs['head']):
                ws[head_list[i] + '1'] = head

        else:
            # default head
            ws['A1'] = '原始词'
            ws['B1'] = '标准词'

        for line in data:
            try:
                ws.append(line)
            except Exception as e:
                print(e)
                assert False, "数据添加错误，错误数据为：" + str(line)

        wb.save(file_name)

    def save_to_txt(self, data: list, file_name: str, **kwargs):
        with open(file_name, 'a', encoding='utf-8') as f:
            for line in data:
                if isinstance(line, list):
                    for item in line[:-1]:
                        f.write(str(item) + "\t")
                    f.write(str(line[-1]) + "\n")

                else:
                    f.write(str(line) + "\n")


class CrimeSpider:
    def __init__(self):
        self.data_path = "./alias.json"
        self.log_path = "./log.txt"
        self.entity_normalization_path = "./alias.xlsx"

    def get_html(self, url):
        '''
        function: 根据url，请求html
        '''
        headers = {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/111.0',
                   'Accept': 'image/avif,image/webp,*/*',
                   'Cookie': 'BAIDUID=88FC0434869B5DDE7F006BEBF4154A4E:SL=0:NR=10:FG=1; BIDUPSID=88FC0434869B5DDEC7D5BDBCAFD29EA9; PSTM=1651401144; __bid_n=1861cb75fc64b2211c4207; FPTOKEN=HtlvyPY3LLL2J0HPOiRoBBZb2AtEczguX2ST2mo3zxz50HhBV8o3mO7MI4OF8BgmUWIiIPlv18uSbw5ztth04mLA9aQ/Fe44dN99UT6A5xrTvleOq+ueQXaDfbXBU6Odt9bfTG0sXUrM52uC+08V2sfv06tBconGFDAEkZ0UnO1iiJGLhglXX+gQ5mVzJinK1+WOL8NzNRX+OglM8UthpXBZEh6FKRHbZ4vXD4AHCnqeWPRDCQJ7Juudqv+LdOmNIEVfi6OiBlxMlgmQHJvp9URgSWpRQTrvC2YGRAlP4Jl2Hl1Otj8R3hNeInxxv7xfui02meVaNjENLo9S13iJQrU5SGSuaWxZu5bITh8T6P1bN4WVxZPmUHe8kzdZX5eH9RE59H2D966feSIh9EJSAA==|jEXc7aciZcQRoAklVrzjh9QkuNxVMm7V9iFxmVnpWy8=|10|f69b240d7776e14ad67dc707447c3dc0; ZFY=:AM1qxkLc:AkiwiqcwWKEVM:B0CQgkPfbx:AqFkJKz4zK48:C; RT="z=1&dm=baidu.com&si=d686e552-a3e5-47eb-8aa4-a174029e9918&ss=lg7x1h61&sl=8&tt=5g6&bcn=https%3A%2F%2Ffclog.baidu.com%2Flog%2Fweirwood%3Ftype%3Dperf&ld=1r31&nu=w62gfu2s&cl=1rwg&ul=1zz8&hd=202m"; ZD_ENTRY=bing; H_PS_PSSID=36544_38127_38396_37862_38173_38289_38217_37919_38314_38383_38285_26350_38422_37881; BCLID=8183660660821604633; BDSFRCVID=bLAOJeC62A8l0d7f5FDseZD17zT3w66TH6_nq-jOHCD6p3BfuN-_EG0PWM8g0KAbPc4wogKKX2OTHIIF_2uxOjjg8UtVJeC6EG0Ptf8g0f5; H_BDCLCKID_SF=JbI8oI--JKvseJj15-A_2-K-5fIX5-CsKCoJ2hcH0KLKbfKmDCc_b6Lz5h7-Xf0DJKvq5KJFKMb1Mx7kKRjEKUuD-GbAKltHa2bd2p5TtUJieCnTDMRNqtv-BNOyKMniynr9-pnYWlQrh459XP68bTkA5bjZKxtq3mkjbPbDfn02eCKuD6uKe5c3eaAs2bO-atoELR7b2n7hJJTRM-o_bICShUFs5f7CB2Q-5KL-0CnShb-6b6oPQbt8W48DtbTnt5kq_fbdJf7_qj6FXU7TQ5KQDP5h-xCLX2TxoUJ6BCnJhhvq-l3jjb-ebPRiXTj9QgbAalQ7tt5W8ncFbT7l5hKpbt-q0x-jLTnhVn0M5DK0HPonHjL-ejOB3H; ab_sr=1.0.1_OWVmNjEyNTA1ZTkyY2RiNGIwNmJjYjNmOGMxODM1YTJhZTc2YTBmNGFiNmU3NTc4MmFiZjNiNjMxMDk2MTljNjg1NTAyZmY4YWIzYzQ2NWVmYTc4ZDFlNzY4ZGEzNzNjNzZlYzJkMzRkNzI2YWFkNjZkOWNkMDczMWJhYzQwNzVjZGVlZTY1MThmYmY3ZGNlMjM0NWRkYmIxNjRhNGYyMQ=='
                   }
        req = urllib.request.Request(url=url, headers=headers)
        res = urllib.request.urlopen(req)
        html = res.read().decode('utf-8')
        return html

    def basicinfo_spider(self, html):
        '''
        function: 基本信息解析
        '''
        selector = etree.HTML(html)
        sen_list = selector.xpath('//div[contains(@class,"lemma-summary") or '
                                  'contains(@class,"lemmaWgt-lemmaSummary")]//text()')
        sen_list_after_filter = [item.strip('\n') for item in sen_list]
        return ''.join(sen_list_after_filter)

    def str_2_list(self, data: str):
        """
        function: 从字符串中抽取出列表
        """
        ans = re.sub(r"、|或|，|等|和", " ", data)
        target = [re.sub(r"\s|\n", "", item) for item in ans.split()]

        return target

    def export_error(self, data, log_info):
        """
        function: 导出错误数据信息
        """
        with open(self.log_path, 'a', encoding='utf-8') as f:
            f.write(log_info + "数据为: " + str(data))
            f.write("\n")

    def strip_null(self, data: list, needed_n=0) -> list:
        """
        function: 去除列表中空白字符
        """
        target = []
        for item in data:
            ans = re.sub(r"\r|\t|\n|\s", "", item)
            if ans:
                target.append(ans)

        # 特殊处理value
        if needed_n:
            value_string = ''.join(data)
            # 保留英文名中的空格
            value_string = re.sub(r" {1,}", "#", value_string)
            # value_string = re.sub(r"\r|\t|\n|\s", "#", value_string)
            # target = value_string.split('#')
            target = re.split(r"\n{1,}|\r{1,}|\t{1,}|\s{1,}|\[[0-9]{1,}\]", value_string)
            target = [item for item in target if item]
            target = [re.sub(r"#{1,}", " ", item) for item in target]

            if len(target) != needed_n:
                # self.export_error(data, "长度不匹配！")
                assert False, "长度不匹配！数据为：" + str(data)
            else:
                return target

        # 普通去除
        else:
            return target

    def is_exist(self, data, target: str):
        """
        function: 检查data中是否有指定值，返回包含该值的项
        example: 症状可能为： 常见症状、症状
        """
        if isinstance(data, list):
            for item in data:
                # if target in item:
                if re.findall(target, item):
                    return item
            return False

        if isinstance(data, str):
            if re.findall(target, data):
                return data
            else:
                return False

    def basicinfo_detail_spider(self, html):
        '''
        function: 详细信息解析
        '''
        selector = etree.HTML(html)
        basicInfo_item_name = selector.xpath('//dt[contains(@class,"basicInfo-item name")]//text()')
        basicInfo_item_name = self.strip_null(basicInfo_item_name)
        basicInfo_item_value = selector.xpath('//dd[contains(@class,"basicInfo-item value")]//text()')
        basicInfo_item_value = self.strip_null(basicInfo_item_value, needed_n=len(basicInfo_item_name))

        basicInfo = {}
        for name, value in zip(basicInfo_item_name, basicInfo_item_value):
            basicInfo[name] = value

        # 处理具体信息
        if '别名' in basicInfo.keys() or '临床称呼' in basicInfo.keys():
            alias_list = []
            if '别名' in basicInfo.keys():
                alias_list.extend(self.str_2_list(basicInfo['别名']))
            if '临床称呼' in basicInfo.keys():
                alias_list.extend(self.str_2_list(basicInfo['临床称呼']))
            basicInfo['别名'] = alias_list
        else:
            basicInfo['别名'] = []

        symptom = self.is_exist(list(basicInfo.keys()), "症状")
        if symptom:
            basicInfo['症状'] = self.str_2_list(basicInfo[symptom])
            del basicInfo[symptom]
        else:
            basicInfo['症状'] = []

        cause = self.is_exist(list(basicInfo.keys()), "病因")
        if cause:
            basicInfo['病因'] = basicInfo[cause]
            del basicInfo[cause]
        else:
            basicInfo['病因'] = []

        return basicInfo

    def get_suburl(self, html):
        """
        function: 针对目标网页在子网页下
        """
        selector = etree.HTML(html)
        sen_list = selector.xpath('//div[contains(@class,"para")]/a/@href')
        # 默认取第一个网页url
        if sen_list:
            return sen_list[0]
        else:
            return ""

    def test_html(self, url):
        """
        function: test url
        """
        html = self.get_html(url)
        if self.basicinfo_spider(html):
            return url, html
        else:
            base_url = "https://baike.baidu.com"
            result_url = self.get_suburl(html)
            if not result_url:
                return url, html
            else:
                result_url = base_url + result_url
                html = self.get_html(result_url)
                return result_url, html

    def spider_main(self, entity: str):
        '''
        function: 抓取主程序
        '''
        url = "https://baike.baidu.com/item/" + urllib.parse.quote(entity)
        try:
            url, html = self.test_html(url)

            data = {}
            # data['url'] = url
            # data['name'] = entity
            # data['desc'] = self.basicinfo_spider(html)

            basic_detail = self.basicinfo_detail_spider(html)
            # data['alias'] = basic_detail['别名']
            # data['symptom'] = basic_detail['症状']
            # data['cause'] = basic_detail['病因']

            # print(data)
            # 这里只使用“别名”数据
            # txt文本每行两列：原实体;别名
            data[entity] = basic_detail['别名']
            with open(self.data_path, "a", encoding='utf-8') as f:
                json.dump(data, f)
                f.write("\n")

        except Exception as e:
            print(e, entity)

    def preprocess_alias(self):
        """
        function: 预处理alias数据
        """
        raw_data_path = self.data_path
        target_data_path = self.entity_normalization_path
        assert not os.path.exists(target_data_path), "文件已存在！"

        # process data
        target_data = []
        with open(raw_data_path, 'r', encoding='utf-8') as f:
            for line in f.readlines():
                data = json.loads(line)
                name = list(data.keys())[0]
                alias = list(data.values())[0]

                if name not in alias:
                    alias.append(name)
                target_alias = '##'.join(alias)
                target_data.append([name, target_alias])

        # append lost diseases
        diseases = get_diseases()
        stored_diseases = [item[0] for item in target_data]
        for disease in diseases:
            if disease not in stored_diseases:
                target_data.append([disease, disease])

        # save data
        saver = SaveFile()
        saver.save_to_xlsx(target_data, target_data_path)


def crawl_main(entities: list, thread_num=60):
    """
    function: 多线程爬取
    """
    # 填充队列
    workQueue = queue.Queue(len(entities))
    for entity in entities:
        workQueue.put(entity)

    threads = []
    for i in range(1, thread_num + 1):
        thread = myThread("Thread-" + str(i), q=workQueue)
        # 开启新线程
        thread.start()
        # 添加新线程到线程列表
        threads.append(thread)

    # 等待所有线程完成
    for thread in threads:
        thread.join()


def get_diseases() -> list:
    """
    function: 读取爬取的实体
    """
    with open("../graph_data/diseases.json", 'r') as f:
        diseases = json.load(f)

    return diseases


if __name__ == '__main__':
    # entities = get_diseases()
    # crawl_main(entities, thread_num=60)
    spider = CrimeSpider()
    # spider.spider_main("感冒")
    spider.preprocess_alias()
